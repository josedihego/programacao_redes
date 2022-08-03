# -*- coding: utf-8 -*-

# 1. Quais usuários acessaram um determinado host

from openpyxl import load_workbook
from datetime import datetime

ip_procurado = input('Informe o IP do host para listar acessos: ')

arquivo_usuarios = load_workbook(filename='planilhas/usuarios.xlsx')

planilha_usuarios = arquivo_usuarios['usuarios']

lista_acessos = [] # lista de acessos ao determinado IP
# lembrar que essa lista começa vazia 

for linha in planilha_usuarios.iter_rows(min_row=2, max_row=31, min_col=1, max_col=3):
    usuario = linha[0].value
    ip = linha[1].value
    data = linha[2].value
    if ip_procurado==ip:
        lista_acessos.append((usuario,data.strftime('%d-%m-%Y')))
        # quando o usuario acessa o IP ip_procurado
        # se coloca o usuario na lista_acessos

if lista_acessos == []:
    print('Nenhum usuário acessou o IP ' + ip_procurado)
else:
    print('Lista dos usuários que acessaram o IP ' + ip_procurado)
    print(lista_acessos)