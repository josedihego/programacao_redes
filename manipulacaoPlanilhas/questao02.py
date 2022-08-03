# -*- coding: utf-8 -*-

# 2. Quais usuários estiveram em um determinado host em um determinado intervalo de tempo

from openpyxl import load_workbook
from datetime import datetime

ip_procurado = input('Informe o IP do host para listar acessos: ')

data_inicial = datetime.strptime(input('Informe a data inicial: '),'%d-%m-%Y')
data_final = datetime.strptime(input('Informe a data final: '),'%d-%m-%Y')

#datetime.strptime(data_proc, '%d-%m-%Y')

arquivo_usuarios = load_workbook(filename='planilhas/usuarios.xlsx')

planilha_usuarios = arquivo_usuarios['usuarios']

lista_acessos = []  # lista de acessos ao determinado IP
# lembrar que essa lista começa vazia

for linha in planilha_usuarios.iter_rows(min_row=2, max_row=31, min_col=1, max_col=3):
    usuario = linha[0].value
    ip = linha[1].value
    data = linha[2].value
    if ip_procurado == ip and data >= data_inicial and data <= data_final:
        lista_acessos.append((usuario, data.strftime('%d-%m-%Y')))
        # quando o usuario acessa o IP ip_procurado
        # se coloca o usuario na lista_acessos

if lista_acessos == []:
    print('Nenhum usuário acessou o IP no intervalo informado')
else:
    print('Lista dos usuários que acessaram o IP no intervalo informado')
    print(lista_acessos)
