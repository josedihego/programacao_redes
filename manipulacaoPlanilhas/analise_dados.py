# -*- coding: utf-8 -*-

# 1. Quais usuários acessaram um determinado host
# 2. Quais usuários estiveram em um determinado host em uma certa data
# 3. Qual a data onde houve uma maior taxa de download?
# 4. Qual a data onde houve uma menor taxa de upload?

from openpyxl import load_workbook
from datetime import datetime


IPs_geral = load_workbook(filename='planilhas/IP.xlsx')
IPs_planilha = IPs_geral['IPs']

for linha in IPs_planilha.iter_rows(min_row=2, max_row=21, min_col=1, max_col=4):
    ip = linha[0].value
    data = linha[1].value
    upload =  linha[2].value
    download = linha[3].value
    #print(data == datetime.strptime('31-07-2020', '%d-%m-%Y'))

# 1. Quais usuários acessaram um determinado host


planilhas_usuario = load_workbook(filename='planilhas/usuarios.xlsx')
planilha_usuario = planilhas_usuario['usuarios']

ip_proc = input('Informe o IP a ser verificado: ')
nenhum_acesso = True
for linha in planilha_usuario.iter_rows(min_row=2, max_row=31, min_col=1, max_col=3):
    usuario = linha[0].value
    ip = linha[1].value
    if(ip == ip_proc):
        nenhum_acesso = False
        print(usuario)

if nenhum_acesso==True:
    print('Nenhum acesso detectado.')


# 2. Quais usuários estiveram em um determinado host em uma certa data

ip_proc = input('Informe o IP a ser verificado: ')
data_proc = input('Informe a data de acesso a ser verificada')
data_formatada = datetime.strptime(data_proc, '%d-%m-%Y')
nenhum_acesso = True
for linha in planilha_usuario.iter_rows(min_row=2, max_row=31, min_col=1, max_col=3):
    usuario = linha[0].value
    ip = linha[1].value
    data = linha[2].value
    if(ip == ip_proc and data == data_formatada):
        nenhum_acesso = False
        print(usuario)

if nenhum_acesso==True:
    print('Nenhum acesso detectado ao IP informado na data informada')