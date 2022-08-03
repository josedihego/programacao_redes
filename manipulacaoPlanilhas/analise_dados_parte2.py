from openpyxl import load_workbook
from datetime import datetime


IPs_geral = load_workbook(filename='../planilhas/IP.xlsx')
IPs_planilha = IPs_geral['IPs']

datas_sem_repeticao = []
lista_taxa_downloads  = []

# 3. Qual a data onde houve uma maior taxa de download?

for linha in IPs_planilha.iter_rows(min_row=2, max_row=21, min_col=1, max_col=4):
    data = linha[1].value
    taxa_download = linha[3].value
    lista_taxa_downloads.append([data,taxa_download])
    if(datas_sem_repeticao.count(data) == 0):
        datas_sem_repeticao.append(data)

taxa_maxima = 0
data_maxima = None

for data in datas_sem_repeticao:
    somatorio_downloads = 0
    for par in lista_taxa_downloads:
        if par[0]== data:
            somatorio_downloads = somatorio_downloads + par[1]

    if somatorio_downloads > taxa_maxima:
        taxa_maxima = somatorio_downloads
        data_maxima = data

print('Na data ', str(data_maxima), ' foi registrada uma taxa máxima de download de ', taxa_maxima)