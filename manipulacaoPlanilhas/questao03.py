# 3. Qual a data onde houve uma maior taxa de download?

# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from datetime import datetime
from functools import total_ordering

@total_ordering
class Consumo:
    def __init__(self, data, taxa_download):
        self.data = data
        self.taxa_download = taxa_download

    def __str__(self):
        return 'data: ' + str(self.data) + ' consumo download: ' + str(self.taxa_download)
    
    def __repr__(self):
         return 'data: ' + str(self.data) + ' consumo download: ' + str(self.taxa_download)

    def adicionar(self, novo_consumo):
        self.taxa_download = self.taxa_download + novo_consumo.taxa_download

    def __eq__(self, outro):
        return self.data == outro.data   
    
    def __lt__(self,outro):
        return self.taxa_download <  outro.taxa_download


class Historico:
    def __init__(self):
        self.lista_comsumo = []

    def adicionar(self, novo_consumo):
        if(self.lista_comsumo.count(novo_consumo) > 0):
            posicao = self.lista_comsumo.index(novo_consumo) 
            consumo_atual = self.lista_comsumo[posicao]
            consumo_atual.adicionar(novo_consumo)
        else:
            self.lista_comsumo.append(novo_consumo)       
    def imprimir(self):
        for c in self.lista_comsumo:
            print (c)

c1 = Consumo(datetime.strptime('24-11-2021 21:06:30', '%d-%m-%Y %H:%M:%S'), 800)
c2 = Consumo(datetime.strptime('24-11-2021 21:06:30', '%d-%m-%Y %H:%M:%S'), 300)

c3 = Consumo(datetime.strptime('25-04-2022 05:09:12', '%d-%m-%Y %H:%M:%S'), 2560)
c4 = Consumo(datetime.strptime('25-04-2022 05:09:12', '%d-%m-%Y %H:%M:%S'), 2400)


h = Historico()
h.adicionar(c1)
h.adicionar(c2)
h.adicionar(c3)
h.adicionar(c4)
h.lista_comsumo.sort(reverse=True)
h.imprimir()


arquivo_IPS = load_workbook(filename='planilhas/IP.xlsx')
planilha = arquivo_IPS['IPs']

for linha in planilha.iter_rows(min_row=2, max_row=21, min_col=1, max_col=4):
    data = linha[1].value
    taxa_download = float(linha[3].value)
    consumo = Consumo(data,taxa_download)
    h.adicionar(consumo)


h.lista_comsumo.sort(reverse=True)
h.imprimir()
